"""Rapor dışa aktarma modülü: JSON, Excel, Markdown ve PDF.

Tüm export'lar disiplin başına raporları içeren bir "bundle" alır:
    bundle = {discipline: (FinalReport, [SpecItem, ...])}
Mekanik ve elektrik bölümleri çıktıların her birinde ayrı sunulur.
"""

from __future__ import annotations

import io
import json
import re
from datetime import datetime

import pandas as pd

from .ai_extractor import FinalReport, SpecItem
from .config import discipline_label
from .utils import logger

# Excel'e yazılamayan kontrol karakterleri (openpyxl IllegalCharacterError'ı önler).
# PDF'ten çıkarılan metinlerde bazen \x00-\x1f aralığında görünmez karakterler kalır.
_ILLEGAL_XLSX_RE = re.compile(r"[\x00-\x08\x0b\x0c\x0e-\x1f]")


def _excel_safe(value):
    """Bir hücre değerinden Excel'e yazılamayan kontrol karakterlerini temizler."""
    if isinstance(value, str):
        return _ILLEGAL_XLSX_RE.sub("", value)
    return value


# --------------------------------------------------------------------------- #
# PDF font kaydı (Türkçe karakter desteği)
# --------------------------------------------------------------------------- #
# Helvetica gibi standart PDF fontları ğ/ş/ı/İ gibi Türkçe karakterleri içermez
# (siyah kutu olarak görünür). Türkçe destekli bir TrueType font kaydedilir;
# bulunamazsa Helvetica'ya düşülür (çıktı yine üretilir).
_TR_FONT_CANDIDATES: list[tuple[str, str | None]] = [
    # (regular, bold | None)
    ("/System/Library/Fonts/Supplemental/Arial Unicode.ttf", None),  # macOS
    ("/Library/Fonts/Arial Unicode.ttf", None),                       # macOS
    (
        "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",            # Linux
        "/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf",
    ),
    ("C:/Windows/Fonts/arial.ttf", "C:/Windows/Fonts/arialbd.ttf"),   # Windows
    ("C:/Windows/Fonts/segoeui.ttf", "C:/Windows/Fonts/segoeuib.ttf"),
]

_FONTS_CACHE: tuple[str, str] | None = None


def _register_pdf_fonts() -> tuple[str, str]:
    """Türkçe destekli (normal, bold) font adlarını döndürür; yoksa Helvetica."""
    global _FONTS_CACHE
    if _FONTS_CACHE is not None:
        return _FONTS_CACHE

    from pathlib import Path

    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont

    base, bold = "Helvetica", "Helvetica-Bold"
    for regular_path, bold_path in _TR_FONT_CANDIDATES:
        if not Path(regular_path).exists():
            continue
        try:
            pdfmetrics.registerFont(TTFont("TRFont", regular_path))
            if bold_path and Path(bold_path).exists():
                pdfmetrics.registerFont(TTFont("TRFont-Bold", bold_path))
                bold_name = "TRFont-Bold"
            else:
                # Ayrı bold dosyası yoksa bold etiketleri de normal font ile çizilir
                bold_name = "TRFont"
            pdfmetrics.registerFontFamily(
                "TRFont", normal="TRFont", bold=bold_name,
                italic="TRFont", boldItalic=bold_name,
            )
            base, bold = "TRFont", bold_name
            logger.info("PDF için Türkçe font kaydedildi: %s", regular_path)
            break
        except Exception as exc:  # noqa: BLE001 - bir sonraki adaya geç
            logger.debug("Font kaydedilemedi (%s): %s", regular_path, exc)

    _FONTS_CACHE = (base, bold)
    return _FONTS_CACHE

# bundle tipi: {discipline: (report, items)}
ReportBundle = "dict[str, tuple[FinalReport, list[SpecItem]]]"

ITEM_COLUMNS = [
    "discipline",
    "category",
    "title",
    "requirement",
    "value_or_limit",
    "related_standard",
    "relevance",
    "source_quote",
    "page_or_section",
    "confidence",
]

_COLUMN_LABELS = {
    "discipline": "Disiplin",
    "category": "Kategori",
    "title": "Başlık",
    "requirement": "Gereklilik",
    "value_or_limit": "Değer / Sınır",
    "related_standard": "Standart",
    "relevance": "İlgili Önem",
    "source_quote": "Kaynak Alıntı",
    "page_or_section": "Sayfa / Bölüm",
    "confidence": "Güven",
}


def items_to_dataframe(items: list[SpecItem]) -> pd.DataFrame:
    """Maddeleri pandas DataFrame'e çevirir (sabit kolon sırasıyla)."""
    if not items:
        return pd.DataFrame(columns=ITEM_COLUMNS)
    rows = []
    for item in items:
        data = item.model_dump()
        data["discipline"] = discipline_label(data.get("discipline", ""))
        rows.append(data)
    df = pd.DataFrame(rows)
    for col in ITEM_COLUMNS:
        if col not in df.columns:
            df[col] = None
    return df[ITEM_COLUMNS]


def _first_title(bundle) -> str:
    """Bundle'daki ilk raporun belge başlığını döndürür."""
    for report, _items in bundle.values():
        if report.document_title:
            return report.document_title
    return "Teknik Şartname"


# --------------------------------------------------------------------------- #
# JSON
# --------------------------------------------------------------------------- #
def export_json(bundle) -> bytes:
    """Disiplin başına rapor + maddeleri JSON bytes olarak döndürür."""
    disciplines_payload = {}
    total_items = 0
    for discipline, (report, items) in bundle.items():
        disciplines_payload[discipline] = {
            "label": discipline_label(discipline),
            "report": report.model_dump(),
            "items": [item.model_dump() for item in items],
            "item_count": len(items),
        }
        total_items += len(items)

    payload = {
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "document_title": _first_title(bundle),
        "disciplines": disciplines_payload,
        "total_item_count": total_items,
    }
    return json.dumps(payload, ensure_ascii=False, indent=2).encode("utf-8")


# --------------------------------------------------------------------------- #
# Excel
# --------------------------------------------------------------------------- #
def export_excel(bundle) -> bytes:
    """Her disiplin için ayrı madde + checklist sayfaları içeren Excel döndürür."""
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    item_widths = [12, 20, 26, 50, 18, 20, 40, 40, 16, 10]

    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        checklist_sheets: set[str] = set()
        for discipline, (report, items) in bundle.items():
            label = discipline_label(discipline)
            df = items_to_dataframe(items).rename(columns=_COLUMN_LABELS)
            df = df.map(_excel_safe)  # kontrol karakterlerini temizle
            items_sheet = f"{label} Maddeler"[:31]
            df.to_excel(writer, sheet_name=items_sheet, index=False)

            checklist_rows = [_excel_safe(c) for c in (report.checklist or [])]
            df_checklist = pd.DataFrame(
                {"#": range(1, len(checklist_rows) + 1), "Kontrol Maddesi": checklist_rows}
            )
            checklist_sheet = f"{label} Checklist"[:31]
            df_checklist.to_excel(writer, sheet_name=checklist_sheet, index=False)
            checklist_sheets.add(checklist_sheet)

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="2F5496")
        wrap = Alignment(vertical="top", wrap_text=True)

        for sheet_name, ws in writer.sheets.items():
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(vertical="center", horizontal="center")
            widths = [6, 70] if sheet_name in checklist_sheets else item_widths
            for col_idx, width in enumerate(widths, start=1):
                ws.column_dimensions[get_column_letter(col_idx)].width = width
            for row in ws.iter_rows(min_row=2):
                for cell in row:
                    cell.alignment = wrap
            ws.freeze_panes = "A2"

    buffer.seek(0)
    return buffer.getvalue()


# --------------------------------------------------------------------------- #
# Markdown
# --------------------------------------------------------------------------- #
def export_markdown(bundle) -> str:
    """Raporu Markdown metni olarak döndürür (disiplin başına bölüm)."""
    lines: list[str] = []
    lines.append("# Şartname Analiz Raporu (Mekanik + Elektrik)")
    lines.append("")
    lines.append(f"**Belge:** {_first_title(bundle)}  ")
    lines.append(f"**Oluşturulma:** {datetime.now().strftime('%Y-%m-%d %H:%M')}  ")
    total = sum(len(items) for _r, items in bundle.values())
    lines.append(f"**Toplam madde:** {total}")
    lines.append("")

    for discipline, (report, items) in bundle.items():
        label = discipline_label(discipline)
        lines.append(f"# {label} Bölümü")
        lines.append("")
        lines.append(f"**{label} madde sayısı:** {len(items)}")
        lines.append("")
        _markdown_section(lines, report, items)

    return "\n".join(lines)


def _markdown_section(lines: list[str], report: FinalReport, items: list[SpecItem]) -> None:
    """Tek disiplin için Markdown bölümünü `lines`'a ekler."""
    lines.append("## Yönetici Özeti")
    lines.append("")
    lines.append(report.executive_summary or "—")
    lines.append("")

    lines.append("## Kategori Bazlı Detaylar")
    lines.append("")
    if report.grouped_details:
        for category, details in report.grouped_details.items():
            lines.append(f"### {category}")
            for detail in details:
                lines.append(f"- {detail}")
            lines.append("")
    else:
        lines.append("_Kategori detayı bulunamadı._")
        lines.append("")

    lines.append("## Detaylı Maddeler")
    lines.append("")
    if items:
        lines.append("| Kategori | Başlık | Gereklilik | Değer/Sınır | Standart | Sayfa | Güven |")
        lines.append("|---|---|---|---|---|---|---|")
        for it in items:
            lines.append(
                "| {cat} | {title} | {req} | {val} | {std} | {pg} | {conf:.2f} |".format(
                    cat=_md_cell(it.category),
                    title=_md_cell(it.title),
                    req=_md_cell(it.requirement),
                    val=_md_cell(it.value_or_limit),
                    std=_md_cell(it.related_standard),
                    pg=_md_cell(it.page_or_section),
                    conf=it.confidence,
                )
            )
        lines.append("")
    else:
        lines.append("_Madde bulunamadı._")
        lines.append("")

    lines.append("## Checklist")
    lines.append("")
    if report.checklist:
        for c in report.checklist:
            lines.append(f"- [ ] {c}")
    else:
        lines.append("_Checklist boş._")
    lines.append("")

    lines.append("## Eksik / Belirsiz Noktalar")
    lines.append("")
    if report.missing_or_unclear_points:
        for m in report.missing_or_unclear_points:
            lines.append(f"- {m}")
    else:
        lines.append("_Belirgin eksik/belirsiz nokta tespit edilmedi._")
    lines.append("")


def _md_cell(value: str | None) -> str:
    """Markdown tablo hücresi için metni güvenli hale getirir."""
    if not value:
        return "—"
    return str(value).replace("|", "\\|").replace("\n", " ").strip()


# --------------------------------------------------------------------------- #
# PDF
# --------------------------------------------------------------------------- #
def export_pdf(bundle) -> bytes:
    """Raporu PDF bytes olarak döndürür (disiplin başına bölüm, reportlab)."""
    try:
        from reportlab.lib import colors
        from reportlab.lib.enums import TA_LEFT
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
        from reportlab.lib.units import mm
        from reportlab.platypus import (
            Paragraph,
            SimpleDocTemplate,
            Spacer,
            Table,
            TableStyle,
        )
    except ImportError as exc:  # pragma: no cover
        raise RuntimeError(
            "reportlab kurulu değil. PDF export için 'pip install reportlab' çalıştırın."
        ) from exc

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        topMargin=18 * mm,
        bottomMargin=18 * mm,
        leftMargin=16 * mm,
        rightMargin=16 * mm,
        title="Şartname Analiz Raporu",
    )

    base_font, bold_font = _register_pdf_fonts()

    styles = getSampleStyleSheet()
    h1 = ParagraphStyle("H1", parent=styles["Title"], fontName=bold_font, fontSize=18, spaceAfter=8)
    h_disc = ParagraphStyle(
        "HDisc", parent=styles["Heading1"], fontName=bold_font, fontSize=15,
        textColor=colors.HexColor("#1F2A44"), spaceBefore=10, spaceAfter=6,
    )
    h2 = ParagraphStyle(
        "H2", parent=styles["Heading2"], fontName=bold_font, fontSize=13,
        textColor=colors.HexColor("#2F5496"),
    )
    body = ParagraphStyle(
        "Body", parent=styles["Normal"], fontName=base_font, fontSize=9, leading=13, alignment=TA_LEFT
    )
    small = ParagraphStyle("Small", parent=styles["Normal"], fontName=base_font, fontSize=7.5, leading=10)

    elements: list = []
    elements.append(Paragraph("Şartname Analiz Raporu (Mekanik + Elektrik)", h1))
    total = sum(len(items) for _r, items in bundle.values())
    elements.append(
        Paragraph(
            f"<b>Belge:</b> {_pdf_text(_first_title(bundle))} &nbsp;&nbsp; "
            f"<b>Tarih:</b> {datetime.now().strftime('%Y-%m-%d %H:%M')} &nbsp;&nbsp; "
            f"<b>Toplam madde:</b> {total}",
            body,
        )
    )
    elements.append(Spacer(1, 8))

    for discipline, (report, items) in bundle.items():
        label = discipline_label(discipline)
        elements.append(Paragraph(f"{label} Bölümü ({len(items)} madde)", h_disc))
        _pdf_section(elements, report, items, colors, mm, Paragraph, Spacer, Table, TableStyle, h2, body, small)

    try:
        doc.build(elements)
    except Exception as exc:
        logger.error("PDF oluşturulamadı: %s", exc)
        raise RuntimeError(f"PDF oluşturulamadı: {exc}") from exc

    buffer.seek(0)
    return buffer.getvalue()


def _pdf_section(elements, report, items, colors, mm, Paragraph, Spacer, Table, TableStyle, h2, body, small):
    """Tek disiplin için PDF bölümünü `elements`'a ekler."""
    elements.append(Paragraph("Yönetici Özeti", h2))
    elements.append(Paragraph(_pdf_text(report.executive_summary) or "—", body))
    elements.append(Spacer(1, 8))

    elements.append(Paragraph("Kategori Bazlı Detaylar", h2))
    if report.grouped_details:
        for category, details in report.grouped_details.items():
            elements.append(Paragraph(f"<b>{_pdf_text(category)}</b>", body))
            for detail in details:
                elements.append(Paragraph(f"• {_pdf_text(detail)}", body))
            elements.append(Spacer(1, 4))
    else:
        elements.append(Paragraph("—", body))
    elements.append(Spacer(1, 6))

    elements.append(Paragraph("Detaylı Maddeler", h2))
    if items:
        header = ["Kategori", "Başlık", "Gereklilik", "Değer", "Standart", "Güven"]
        table_data = [[Paragraph(f"<b>{h}</b>", small) for h in header]]
        for it in items:
            table_data.append(
                [
                    Paragraph(_pdf_text(it.category), small),
                    Paragraph(_pdf_text(it.title), small),
                    Paragraph(_pdf_text(it.requirement), small),
                    Paragraph(_pdf_text(it.value_or_limit) or "—", small),
                    Paragraph(_pdf_text(it.related_standard) or "—", small),
                    Paragraph(f"{it.confidence:.2f}", small),
                ]
            )
        table = Table(
            table_data,
            colWidths=[28 * mm, 30 * mm, 55 * mm, 22 * mm, 25 * mm, 12 * mm],
            repeatRows=1,
        )
        table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2F5496")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                    ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#B0B0B0")),
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                    ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F2F4F8")]),
                    ("TOPPADDING", (0, 0), (-1, -1), 3),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
                ]
            )
        )
        elements.append(table)
    else:
        elements.append(Paragraph("Madde bulunamadı.", body))
    elements.append(Spacer(1, 8))

    elements.append(Paragraph("Checklist", h2))
    if report.checklist:
        for c in report.checklist:
            elements.append(Paragraph(f"☐ {_pdf_text(c)}", body))
    else:
        elements.append(Paragraph("—", body))
    elements.append(Spacer(1, 8))

    elements.append(Paragraph("Eksik / Belirsiz Noktalar", h2))
    if report.missing_or_unclear_points:
        for m in report.missing_or_unclear_points:
            elements.append(Paragraph(f"• {_pdf_text(m)}", body))
    else:
        elements.append(Paragraph("Belirgin eksik/belirsiz nokta tespit edilmedi.", body))
    elements.append(Spacer(1, 10))


def _pdf_text(value: str | None) -> str:
    """reportlab Paragraph için XML kaçışı uygular."""
    if not value:
        return ""
    return (
        str(value)
        .replace("&", "&amp;")
        .replace("<", "&lt;")
        .replace(">", "&gt;")
        .strip()
    )
